#!/usr/bin/env python3
"""
Image SERP 收錄批量檢測工具 (SerpAPI 版)
==========================================
透過 SerpAPI 串接 Google 圖片搜尋，批量驗證網頁 URL 的圖片是否被收錄。

使用方式:
  python image_serp_checker.py -i urls.txt -o results.xlsx --key YOUR_SERPAPI_KEY
  python image_serp_checker.py -i urls.txt -o report.html --key YOUR_SERPAPI_KEY
  python image_serp_checker.py -i urls.csv -o results.csv -c 5 -d 1.0
  python image_serp_checker.py -i urls.txt -o results.xlsx              # 使用環境變數 SERPAPI_KEY

環境變數:
  SERPAPI_KEY=xxxxxxxx    設定後可省略 --key 參數

SerpAPI 註冊: https://serpapi.com (免費 100 次/月)

依賴安裝:
  pip install requests openpyxl tqdm
"""

import argparse
import csv
import json
import os
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

try:
    import requests
except ImportError:
    print("❌ 請先安裝 requests:  pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False


# ═══════════════════════════════════════════════════════
#  設定
# ═══════════════════════════════════════════════════════
SERPAPI_BASE = "https://serpapi.com/search.json"
DEFAULT_CONCURRENCY = 3
DEFAULT_DELAY = 1.0       # SerpAPI 官方無嚴格速率限制，但建議留間隔

# 控制旗標
pause_event = threading.Event()
pause_event.set()
abort_flag = threading.Event()


# ═══════════════════════════════════════════════════════
#  工具函式
# ═══════════════════════════════════════════════════════
def load_urls(filepath: str) -> list[str]:
    """從 TXT / CSV 載入 URL"""
    path = Path(filepath)
    if not path.exists():
        print(f"  ❌ 找不到檔案: {filepath}")
        sys.exit(1)

    text = path.read_text(encoding="utf-8-sig")
    urls = []

    if path.suffix.lower() == ".csv":
        for row in csv.reader(text.splitlines()):
            for cell in row:
                cell = cell.strip()
                if cell.startswith(("http://", "https://")):
                    urls.append(cell)
    else:
        for line in text.splitlines():
            line = line.strip()
            if line.startswith(("http://", "https://")):
                urls.append(line)

    # 去重 (保留順序)
    seen = set()
    return [u for u in urls if not (u in seen or seen.add(u))]


def domain_of(url: str) -> str:
    try:
        return urlparse(url).hostname or ""
    except Exception:
        return ""


def path_of(url: str) -> str:
    try:
        p = urlparse(url)
        return p.path + ("?" + p.query if p.query else "")
    except Exception:
        return "/"


def url_match(candidate: str, target_url: str, target_domain: str) -> bool:
    """判斷搜尋結果中的連結是否匹配目標 URL"""
    if not candidate:
        return False
    # 完全匹配（忽略結尾斜線）
    c = candidate.rstrip("/")
    t = target_url.rstrip("/")
    if c == t:
        return True
    # 同域名 + 路徑匹配
    try:
        cp = urlparse(candidate)
        tp = urlparse(target_url)
        if cp.hostname == tp.hostname and cp.path.rstrip("/") == tp.path.rstrip("/"):
            return True
    except Exception:
        pass
    # 域名包含（寬鬆）
    if target_domain and target_domain in candidate:
        return True
    return False


# ═══════════════════════════════════════════════════════
#  SerpAPI 查詢核心
# ═══════════════════════════════════════════════════════
class SerpAPIChecker:
    """
    對每個 URL 執行查詢:
      1. engine=google_images  →  判斷圖片是否被收錄 + 排名
      2. engine=google         →  判斷頁面是否被索引 (可選)

    每個 URL 消耗 1~2 次 SerpAPI credit。
    加 --skip-index 跳過步驟 2，只消耗 1 credit/URL。
    """

    def __init__(self, api_key: str, hl: str = "zh-TW", gl: str = "tw",
                 skip_index: bool = False):
        self.api_key = api_key
        self.hl = hl
        self.gl = gl
        self.skip_index = skip_index
        self.session = requests.Session()

    def _call(self, params: dict) -> dict:
        """呼叫 SerpAPI"""
        params["api_key"] = self.api_key
        params["output"] = "json"
        r = self.session.get(SERPAPI_BASE, params=params, timeout=60)
        r.raise_for_status()
        return r.json()

    def check(self, url: str) -> dict:
        """檢查單一 URL 的圖片收錄狀態"""
        domain = domain_of(url)
        path = path_of(url)
        query = f"site:{domain}{path}"

        result = {
            "url": url,
            "domain": domain,
            "status": "not_found",
            "indexed": False,
            "image_found": False,
            "image_count": 0,
            "best_position": None,
            "image_titles": "",
            "image_urls": "",
            "source_urls": "",
            "snippet": "",
        }

        try:
            # ──── 1. Google 圖片搜尋 ────
            # SerpAPI engine=google_images
            # 回傳 images_results[] 每筆含:
            #   position, title, link (來源頁), source (網站名),
            #   original (圖片原始URL), original_width, original_height
            data = self._call({
                "engine": "google_images",
                "q": query,
                "hl": self.hl,
                "gl": self.gl,
                "ijn": "0",
            })

            images = data.get("images_results", [])
            total_reported = data.get("search_information", {}).get("total_results", 0)

            matched = []
            for img in images:
                link = img.get("link", "")           # 圖片來源頁面 URL
                source_name = img.get("source", "")  # 網站名稱
                if url_match(link, url, domain) or (domain and domain in source_name):
                    matched.append({
                        "position": img.get("position"),
                        "title": img.get("title", ""),
                        "image_url": img.get("original", ""),
                        "source_url": link,
                        "width": img.get("original_width"),
                        "height": img.get("original_height"),
                    })

            if matched:
                result["image_found"] = True
                result["image_count"] = len(matched)
                result["best_position"] = matched[0]["position"]
                result["image_titles"] = " | ".join(m["title"] for m in matched[:5])
                result["image_urls"] = " | ".join(m["image_url"] for m in matched[:5])
                result["source_urls"] = " | ".join(m["source_url"] for m in matched[:5])
                result["snippet"] = (
                    f"✓ 找到 {len(matched)} 張圖片，"
                    f"最佳排名第 {matched[0]['position']} 名"
                )
                result["status"] = "found"
            else:
                result["snippet"] = (
                    f"搜尋 \"{query}\" 回傳 {len(images)} 張圖片，"
                    f"未匹配到此 URL (Google 約 {total_reported} 筆)"
                )

            # ──── 2. 一般搜尋確認頁面索引 (可選) ────
            if not self.skip_index:
                data2 = self._call({
                    "engine": "google",
                    "q": query,
                    "hl": self.hl,
                    "gl": self.gl,
                    "num": 10,
                })
                for res in data2.get("organic_results", []):
                    if url_match(res.get("link", ""), url, domain):
                        result["indexed"] = True
                        break

            if result["image_found"]:
                result["indexed"] = True

        except requests.exceptions.HTTPError as e:
            code = e.response.status_code if e.response else 0
            if code == 401:
                result["status"] = "error"
                result["snippet"] = "❌ API Key 無效"
            elif code == 429:
                result["status"] = "error"
                result["snippet"] = "❌ API 額度用盡"
            else:
                result["status"] = "error"
                result["snippet"] = f"HTTP {code}: {str(e)[:120]}"
        except requests.exceptions.RequestException as e:
            result["status"] = "error"
            result["snippet"] = f"網路錯誤: {str(e)[:120]}"
        except Exception as e:
            result["status"] = "error"
            result["snippet"] = f"錯誤: {str(e)[:120]}"

        return result


# ═══════════════════════════════════════════════════════
#  進度條
# ═══════════════════════════════════════════════════════
class Progress:
    def __init__(self, total: int):
        self.total = total
        self.done = 0
        self.found = 0
        self.not_found = 0
        self.errors = 0
        self.lock = threading.Lock()
        if HAS_TQDM:
            self.bar = tqdm(
                total=total, desc="🔍 查詢中",
                bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}] {postfix}",
                ncols=95,
            )
        else:
            self.bar = None

    def update(self, r: dict):
        with self.lock:
            self.done += 1
            s = r.get("status", "error")
            if s == "found":    self.found += 1
            elif s == "error":  self.errors += 1
            else:               self.not_found += 1

            tag = f"✓{self.found}  ✗{self.not_found}  ⚠{self.errors}"
            if self.bar:
                self.bar.set_postfix_str(tag)
                self.bar.update(1)
            else:
                pct = self.done / self.total * 100
                n = 30
                filled = int(n * self.done / self.total)
                bar = "█" * filled + "░" * (n - filled)
                print(f"\r  [{bar}] {pct:5.1f}%  ({self.done}/{self.total})  {tag}",
                      end="", flush=True)

    def close(self):
        if self.bar:
            self.bar.close()
        else:
            print()


# ═══════════════════════════════════════════════════════
#  輸出: CSV
# ═══════════════════════════════════════════════════════
HEADERS = [
    "序號", "URL", "域名", "圖片收錄", "頁面索引",
    "圖片數量", "最佳排名", "圖片標題", "圖片網址", "來源頁面", "摘要",
]

def save_csv(results: list[dict], filepath: str):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for i, r in enumerate(results, 1):
            w.writerow([
                i, r["url"], r["domain"],
                "是" if r["image_found"] else "否",
                "是" if r["indexed"] else "否",
                r["image_count"],
                r["best_position"] or "",
                r["image_titles"],
                r["image_urls"],
                r["source_urls"],
                r["snippet"],
            ])
    print(f"  ✅ CSV 已儲存: {filepath}")


# ═══════════════════════════════════════════════════════
#  輸出: Excel (含格式化 + 多 Sheet)
# ═══════════════════════════════════════════════════════
def save_excel(results: list[dict], filepath: str):
    if not HAS_OPENPYXL:
        print("  ⚠️  未安裝 openpyxl，自動改輸出 CSV")
        save_csv(results, filepath.replace(".xlsx", ".csv"))
        return

    wb = Workbook()

    # ── Sheet 1: 明細 ──
    ws = wb.active
    ws.title = "查詢結果"

    hdr_font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align

    green_fill = PatternFill(start_color="E6FFF2", fill_type="solid")
    red_fill   = PatternFill(start_color="FFF0F0", fill_type="solid")
    amber_fill = PatternFill(start_color="FFFBE6", fill_type="solid")
    green_font = Font(name="Microsoft JhengHei", color="0D7C3E", bold=True)
    red_font   = Font(name="Microsoft JhengHei", color="CC3333", bold=True)
    amber_font = Font(name="Microsoft JhengHei", color="B45309", bold=True)
    center     = Alignment(horizontal="center")

    for i, r in enumerate(results, 1):
        row = i + 1
        vals = [
            i, r["url"], r["domain"],
            "是" if r["image_found"] else "否",
            "是" if r["indexed"] else "否",
            r["image_count"],
            r["best_position"] or "",
            r["image_titles"],
            r["image_urls"],
            r["source_urls"],
            r["snippet"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            if c in (1, 4, 5, 6, 7):
                cell.alignment = center

        fill = green_fill if r["status"] == "found" else amber_fill if r["status"] == "error" else red_fill
        font = green_font if r["status"] == "found" else amber_font if r["status"] == "error" else red_font
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=4).font = font
        ws.cell(row=row, column=5).font = font

    widths = [6, 55, 22, 10, 10, 10, 10, 40, 45, 45, 45]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(results) + 1}"

    # ── Sheet 2: 統計摘要 ──
    ws2 = wb.create_sheet("統計摘要")
    total  = len(results)
    found  = sum(1 for r in results if r["image_found"])
    nfound = sum(1 for r in results if not r["image_found"] and r["status"] != "error")
    errors = sum(1 for r in results if r["status"] == "error")
    indexed = sum(1 for r in results if r["indexed"])
    rate = f"{found / total * 100:.1f}%" if total else "N/A"

    bold = Font(name="Microsoft JhengHei", bold=True, size=11)
    summary = [
        ("指標", "數值"),
        ("查詢總數", total),
        ("圖片已收錄", found),
        ("圖片未收錄", nfound),
        ("查詢錯誤", errors),
        ("頁面已索引", indexed),
        ("圖片收錄率", rate),
        ("查詢引擎", "SerpAPI → Google Images"),
        ("查詢時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ws2.cell(row=i, column=1, value=k).font = bold
        ws2.cell(row=i, column=2, value=v)
        if i == 1:
            ws2.cell(row=i, column=2).font = bold
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 30

    # ── Sheet 3: 域名統計 ──
    ws3 = wb.create_sheet("域名統計")
    domain_stats = {}
    for r in results:
        d = r["domain"]
        if d not in domain_stats:
            domain_stats[d] = {"total": 0, "found": 0, "not_found": 0, "error": 0}
        domain_stats[d]["total"] += 1
        if r["status"] == "found":      domain_stats[d]["found"] += 1
        elif r["status"] == "error":    domain_stats[d]["error"] += 1
        else:                            domain_stats[d]["not_found"] += 1

    ws3_headers = ["域名", "查詢數", "已收錄", "未收錄", "錯誤", "收錄率"]
    for c, h in enumerate(ws3_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align

    for i, (d, s) in enumerate(sorted(domain_stats.items(), key=lambda x: -x[1]["total"]), 2):
        ws3.cell(row=i, column=1, value=d)
        ws3.cell(row=i, column=2, value=s["total"]).alignment = center
        ws3.cell(row=i, column=3, value=s["found"]).alignment = center
        ws3.cell(row=i, column=4, value=s["not_found"]).alignment = center
        ws3.cell(row=i, column=5, value=s["error"]).alignment = center
        dr = f"{s['found']/s['total']*100:.1f}%" if s["total"] else "N/A"
        ws3.cell(row=i, column=6, value=dr).alignment = center

    ws3.column_dimensions["A"].width = 30
    for c in "BCDEF":
        ws3.column_dimensions[c].width = 12

    wb.save(filepath)
    print(f"  ✅ Excel 已儲存: {filepath}")


# ═══════════════════════════════════════════════════════
#  輸出: 互動式 HTML 報表
# ═══════════════════════════════════════════════════════
def save_html(results: list[dict], filepath: str):
    """產生互動式 HTML 報表 (含篩選、排序、搜尋、圖表、CSV 匯出)"""

    total   = len(results)
    found   = sum(1 for r in results if r["image_found"])
    nfound  = sum(1 for r in results if not r["image_found"] and r["status"] != "error")
    errors  = sum(1 for r in results if r["status"] == "error")
    indexed = sum(1 for r in results if r["indexed"])
    rate    = round(found / total * 100, 1) if total else 0

    # 域名統計
    domain_stats = {}
    for r in results:
        d = r["domain"]
        if d not in domain_stats:
            domain_stats[d] = {"total": 0, "found": 0}
        domain_stats[d]["total"] += 1
        if r["status"] == "found":
            domain_stats[d]["found"] += 1
    top_domains = sorted(domain_stats.items(), key=lambda x: -x[1]["total"])[:20]

    # JSON 資料 (嵌入 HTML)
    rows_json = json.dumps([{
        "i": i + 1,
        "url": r["url"],
        "domain": r["domain"],
        "status": r["status"],
        "indexed": r["indexed"],
        "image_found": r["image_found"],
        "image_count": r["image_count"],
        "best_position": r["best_position"],
        "image_titles": r["image_titles"],
        "image_urls": r["image_urls"],
        "source_urls": r["source_urls"],
        "snippet": r["snippet"],
    } for i, r in enumerate(results)], ensure_ascii=False)

    domains_json = json.dumps([
        {"domain": d, "total": s["total"], "found": s["found"]}
        for d, s in top_domains
    ], ensure_ascii=False)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html = f'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Image SERP 收錄檢測報表</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;500;700&family=JetBrains+Mono:wght@400;500&display=swap');

* {{ margin: 0; padding: 0; box-sizing: border-box; }}
:root {{
  --bg: #0c0c14; --bg2: #13131e; --bg3: #1a1a2a;
  --border: #1e1e30; --border2: #2a2a40;
  --text: #d8d8ec; --text2: #9090b0; --text3: #555570;
  --green: #34d399; --green-bg: #0d3320;
  --red: #f87171; --red-bg: #3b1c1c;
  --amber: #fbbf24; --amber-bg: #3b2e1c;
  --blue: #4ea8de; --purple: #7c3aed;
  --radius: 12px;
}}
body {{
  font-family: 'Noto Sans TC', sans-serif;
  background: var(--bg); color: var(--text);
  line-height: 1.6; min-height: 100vh;
}}
.container {{ max-width: 1400px; margin: 0 auto; padding: 24px 28px; }}

/* Header */
.header {{
  display: flex; align-items: center; justify-content: space-between;
  padding: 20px 0 28px; border-bottom: 1px solid var(--border);
  margin-bottom: 28px; flex-wrap: wrap; gap: 16px;
}}
.header-left {{ display: flex; align-items: center; gap: 14px; }}
.logo {{
  width: 40px; height: 40px; border-radius: 10px;
  background: linear-gradient(135deg, var(--blue), var(--purple));
  display: flex; align-items: center; justify-content: center; font-size: 20px;
}}
.header h1 {{ font-size: 20px; font-weight: 700; letter-spacing: -0.02em; }}
.header .sub {{ font-size: 12px; color: var(--text3); margin-top: 2px; }}
.header-right {{ display: flex; gap: 8px; flex-wrap: wrap; }}

/* Buttons */
.btn {{
  padding: 8px 16px; border-radius: 8px; border: 1px solid var(--border2);
  background: var(--bg2); color: var(--text2); font-size: 13px; font-weight: 500;
  cursor: pointer; transition: all .15s; font-family: inherit;
}}
.btn:hover {{ border-color: var(--blue); color: var(--blue); }}
.btn.active {{ background: #1a2840; border-color: var(--blue); color: var(--blue); }}
.btn-primary {{
  background: linear-gradient(135deg, var(--blue), var(--purple));
  border: none; color: #fff;
}}

/* Stats Cards */
.stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 14px; margin-bottom: 28px; }}
.stat-card {{
  background: var(--bg2); border: 1px solid var(--border); border-radius: var(--radius);
  padding: 18px 20px; text-align: center;
}}
.stat-value {{
  font-size: 28px; font-weight: 700;
  font-family: 'JetBrains Mono', monospace; line-height: 1.2;
}}
.stat-label {{ font-size: 12px; color: var(--text3); margin-top: 4px; font-weight: 500; }}
.stat-value.green {{ color: var(--green); }}
.stat-value.red {{ color: var(--red); }}
.stat-value.amber {{ color: var(--amber); }}
.stat-value.blue {{ color: var(--blue); }}
.stat-value.purple {{ color: var(--purple); }}

/* Charts Area */
.charts {{ display: grid; grid-template-columns: 300px 1fr; gap: 18px; margin-bottom: 28px; }}
@media (max-width: 800px) {{ .charts {{ grid-template-columns: 1fr; }} }}
.chart-card {{
  background: var(--bg2); border: 1px solid var(--border); border-radius: var(--radius);
  padding: 20px;
}}
.chart-title {{ font-size: 13px; font-weight: 600; color: var(--text2); margin-bottom: 16px; }}

/* Donut Chart */
.donut-wrap {{ display: flex; align-items: center; justify-content: center; flex-direction: column; }}
.donut-center {{ font-size: 26px; font-weight: 700; font-family: 'JetBrains Mono', monospace; }}
.donut-legend {{ display: flex; gap: 18px; margin-top: 16px; font-size: 12px; color: var(--text2); }}
.donut-legend span {{ display: flex; align-items: center; gap: 5px; }}
.dot {{ width: 8px; height: 8px; border-radius: 50%; display: inline-block; }}

/* Bar Chart */
.bar-row {{ display: flex; align-items: center; gap: 10px; margin-bottom: 8px; font-size: 12px; }}
.bar-label {{ width: 140px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; color: var(--text2); text-align: right; }}
.bar-track {{ flex: 1; height: 20px; background: var(--bg3); border-radius: 4px; overflow: hidden; position: relative; }}
.bar-fill {{ height: 100%; border-radius: 4px; transition: width .5s; }}
.bar-fill.green {{ background: var(--green); }}
.bar-fill.red {{ background: rgba(248,113,113,0.4); }}
.bar-val {{ font-family: 'JetBrains Mono', monospace; width: 60px; color: var(--text3); font-size: 11px; }}

/* Controls */
.controls {{
  display: flex; align-items: center; gap: 10px; margin-bottom: 16px;
  flex-wrap: wrap;
}}
.search-box {{
  flex: 1; min-width: 200px; padding: 8px 14px; border-radius: 8px;
  border: 1px solid var(--border2); background: var(--bg3); color: var(--text);
  font-size: 13px; font-family: inherit; outline: none;
}}
.search-box:focus {{ border-color: var(--blue); }}
.filter-group {{ display: flex; gap: 4px; }}
.filter-btn {{
  padding: 6px 12px; border-radius: 6px; border: 1px solid var(--border);
  background: transparent; color: var(--text3); font-size: 12px;
  cursor: pointer; font-family: inherit; font-weight: 500; transition: all .15s;
}}
.filter-btn:hover {{ border-color: var(--blue); color: var(--blue); }}
.filter-btn.active {{ background: #1a2840; border-color: var(--blue); color: var(--blue); }}

/* Table */
.table-wrap {{
  background: var(--bg2); border: 1px solid var(--border); border-radius: var(--radius);
  overflow: hidden;
}}
table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
thead th {{
  padding: 12px 14px; text-align: left; font-size: 11px; font-weight: 600;
  color: var(--text3); text-transform: uppercase; letter-spacing: .05em;
  background: var(--bg3); border-bottom: 1px solid var(--border);
  cursor: pointer; user-select: none; white-space: nowrap;
  position: sticky; top: 0; z-index: 10;
}}
thead th:hover {{ color: var(--blue); }}
thead th .sort-arrow {{ font-size: 10px; margin-left: 4px; opacity: .4; }}
thead th.sorted .sort-arrow {{ opacity: 1; color: var(--blue); }}
tbody {{ max-height: 600px; }}
.table-scroll {{ max-height: 620px; overflow-y: auto; }}
.table-scroll::-webkit-scrollbar {{ width: 5px; }}
.table-scroll::-webkit-scrollbar-thumb {{ background: var(--border2); border-radius: 3px; }}
tbody tr {{ border-bottom: 1px solid #15151f; transition: background .1s; }}
tbody tr:hover {{ background: #16162080; }}
td {{ padding: 10px 14px; vertical-align: middle; }}
td.mono {{ font-family: 'JetBrains Mono', monospace; font-size: 12px; }}
td.center {{ text-align: center; }}
td.url-cell {{
  max-width: 320px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
  font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--text2);
}}
td a {{ color: var(--blue); text-decoration: none; }}
td a:hover {{ text-decoration: underline; }}

/* Badges */
.badge {{
  display: inline-block; padding: 3px 10px; border-radius: 6px;
  font-size: 11px; font-weight: 600; white-space: nowrap;
}}
.badge.found {{ background: var(--green-bg); color: var(--green); }}
.badge.not_found {{ background: var(--red-bg); color: var(--red); }}
.badge.error {{ background: var(--amber-bg); color: var(--amber); }}
.yes {{ color: var(--green); font-weight: 600; }}
.no {{ color: var(--red); font-weight: 600; }}

/* Footer */
.footer {{
  margin-top: 24px; padding: 16px 0; border-top: 1px solid var(--border);
  font-size: 11px; color: var(--text3); display: flex; justify-content: space-between;
  flex-wrap: wrap; gap: 8px;
}}
.count-info {{ font-size: 12px; color: var(--text3); margin-top: 10px; }}
</style>
</head>
<body>
<div class="container">

  <!-- Header -->
  <div class="header">
    <div class="header-left">
      <div class="logo">🔍</div>
      <div>
        <h1>Image SERP 收錄檢測報表</h1>
        <div class="sub">透過 SerpAPI × Google 圖片搜尋驗證 · {now}</div>
      </div>
    </div>
    <div class="header-right">
      <button class="btn" onclick="exportCSV()">📥 匯出 CSV</button>
      <button class="btn" onclick="window.print()">🖨️ 列印</button>
    </div>
  </div>

  <!-- Stats -->
  <div class="stats">
    <div class="stat-card">
      <div class="stat-value blue">{total}</div>
      <div class="stat-label">查詢總數</div>
    </div>
    <div class="stat-card">
      <div class="stat-value green">{found}</div>
      <div class="stat-label">圖片已收錄</div>
    </div>
    <div class="stat-card">
      <div class="stat-value red">{nfound}</div>
      <div class="stat-label">圖片未收錄</div>
    </div>
    <div class="stat-card">
      <div class="stat-value amber">{errors}</div>
      <div class="stat-label">查詢錯誤</div>
    </div>
    <div class="stat-card">
      <div class="stat-value purple">{indexed}</div>
      <div class="stat-label">頁面已索引</div>
    </div>
    <div class="stat-card">
      <div class="stat-value green">{rate}%</div>
      <div class="stat-label">圖片收錄率</div>
    </div>
  </div>

  <!-- Charts -->
  <div class="charts">
    <div class="chart-card">
      <div class="chart-title">收錄分佈</div>
      <div class="donut-wrap">
        <svg id="donut" width="180" height="180" viewBox="0 0 180 180"></svg>
        <div class="donut-center" id="donutCenter">{rate}%</div>
        <div class="donut-legend">
          <span><i class="dot" style="background:var(--green)"></i> 已收錄 {found}</span>
          <span><i class="dot" style="background:var(--red)"></i> 未收錄 {nfound}</span>
          <span><i class="dot" style="background:var(--amber)"></i> 錯誤 {errors}</span>
        </div>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-title">域名收錄概覽 (Top 20)</div>
      <div id="domainBars" style="max-height:340px;overflow-y:auto"></div>
    </div>
  </div>

  <!-- Controls -->
  <div class="controls">
    <input type="text" class="search-box" id="search" placeholder="🔎 搜尋 URL 或域名..." oninput="applyFilters()">
    <div class="filter-group">
      <button class="filter-btn active" data-filter="all" onclick="setFilter(this)">全部</button>
      <button class="filter-btn" data-filter="found" onclick="setFilter(this)">✓ 已收錄</button>
      <button class="filter-btn" data-filter="not_found" onclick="setFilter(this)">✗ 未收錄</button>
      <button class="filter-btn" data-filter="error" onclick="setFilter(this)">⚠ 錯誤</button>
    </div>
  </div>

  <!-- Table -->
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th onclick="sortBy('i')" style="width:50px"># <span class="sort-arrow">▲</span></th>
          <th onclick="sortBy('url')">URL <span class="sort-arrow">▲</span></th>
          <th onclick="sortBy('domain')" style="width:140px">域名 <span class="sort-arrow">▲</span></th>
          <th onclick="sortBy('image_found')" style="width:90px;text-align:center">圖片收錄 <span class="sort-arrow">▲</span></th>
          <th onclick="sortBy('indexed')" style="width:90px;text-align:center">頁面索引 <span class="sort-arrow">▲</span></th>
          <th onclick="sortBy('image_count')" style="width:80px;text-align:center">圖片數 <span class="sort-arrow">▲</span></th>
          <th onclick="sortBy('best_position')" style="width:80px;text-align:center">最佳排名 <span class="sort-arrow">▲</span></th>
          <th style="min-width:200px">摘要</th>
        </tr>
      </thead>
    </table>
    <div class="table-scroll">
      <table>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </div>
  <div class="count-info" id="countInfo"></div>

  <!-- Footer -->
  <div class="footer">
    <span>Image SERP 收錄批量檢測工具 · SerpAPI × Google Images</span>
    <span>報表產生時間: {now}</span>
  </div>
</div>

<script>
// ─── Data ───
const DATA = {rows_json};
const DOMAINS = {domains_json};

let filtered = [...DATA];
let currentFilter = 'all';
let sortKey = 'i';
let sortAsc = true;

// ─── Donut Chart ───
function drawDonut() {{
  const svg = document.getElementById('donut');
  const vals = [{found}, {nfound}, {errors}];
  const colors = ['var(--green)', 'var(--red)', 'var(--amber)'];
  const total = vals.reduce((a, b) => a + b, 0) || 1;
  const cx = 90, cy = 90, r = 70, sw = 18;
  let cumAngle = -90;

  svg.innerHTML = '';
  const circ = 2 * Math.PI * r;
  vals.forEach((v, i) => {{
    if (v === 0) return;
    const pct = v / total;
    const dash = circ * pct;
    const gap = circ - dash;
    const rotation = cumAngle;
    cumAngle += pct * 360;
    const circle = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
    circle.setAttribute('cx', cx); circle.setAttribute('cy', cy);
    circle.setAttribute('r', r); circle.setAttribute('fill', 'none');
    circle.setAttribute('stroke', colors[i]); circle.setAttribute('stroke-width', sw);
    circle.setAttribute('stroke-dasharray', `${{dash}} ${{gap}}`);
    circle.setAttribute('transform', `rotate(${{rotation}} ${{cx}} ${{cy}})`);
    circle.style.transition = 'stroke-dasharray .6s';
    svg.appendChild(circle);
  }});
}}

// ─── Domain Bars ───
function drawDomainBars() {{
  const el = document.getElementById('domainBars');
  const maxTotal = Math.max(...DOMAINS.map(d => d.total), 1);
  el.innerHTML = DOMAINS.map(d => {{
    const foundPct = (d.found / maxTotal * 100).toFixed(1);
    const nfPct = ((d.total - d.found) / maxTotal * 100).toFixed(1);
    const rate = (d.found / d.total * 100).toFixed(1);
    return `<div class="bar-row">
      <div class="bar-label" title="${{d.domain}}">${{d.domain}}</div>
      <div class="bar-track">
        <div class="bar-fill green" style="width:${{foundPct}}%;position:absolute"></div>
        <div class="bar-fill red" style="width:${{Number(foundPct)+Number(nfPct)}}%"></div>
      </div>
      <div class="bar-val">${{d.found}}/${{d.total}} (${{rate}}%)</div>
    </div>`;
  }}).join('');
}}

// ─── Table ───
function renderTable() {{
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = filtered.map(r => `<tr>
    <td class="mono center">${{r.i}}</td>
    <td class="url-cell" title="${{esc(r.url)}}"><a href="${{esc(r.url)}}" target="_blank">${{esc(r.url)}}</a></td>
    <td style="font-size:12px;color:var(--text2)">${{esc(r.domain)}}</td>
    <td class="center"><span class="badge ${{r.status}}">${{r.image_found ? '✓ 已收錄' : r.status==='error' ? '⚠ 錯誤' : '✗ 未收錄'}}</span></td>
    <td class="center">${{r.indexed ? '<span class=yes>是</span>' : '<span class=no>否</span>'}}</td>
    <td class="mono center">${{r.image_count}}</td>
    <td class="mono center">${{r.best_position || '—'}}</td>
    <td style="font-size:12px;color:var(--text2);max-width:280px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${{esc(r.snippet)}}">${{esc(r.snippet)}}</td>
  </tr>`).join('');
  document.getElementById('countInfo').textContent = `顯示 ${{filtered.length}} / ${{DATA.length}} 筆`;
}}

function esc(s) {{ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/"/g,'&quot;'); }}

// ─── Filter ───
function setFilter(btn) {{
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  currentFilter = btn.dataset.filter;
  applyFilters();
}}

function applyFilters() {{
  const q = document.getElementById('search').value.toLowerCase();
  filtered = DATA.filter(r => {{
    if (currentFilter !== 'all' && r.status !== currentFilter) return false;
    if (q && !r.url.toLowerCase().includes(q) && !r.domain.toLowerCase().includes(q)) return false;
    return true;
  }});
  doSort();
  renderTable();
}}

// ─── Sort ───
function sortBy(key) {{
  if (sortKey === key) sortAsc = !sortAsc;
  else {{ sortKey = key; sortAsc = true; }}
  document.querySelectorAll('thead th').forEach(th => th.classList.remove('sorted'));
  doSort();
  renderTable();
}}

function doSort() {{
  filtered.sort((a, b) => {{
    let va = a[sortKey], vb = b[sortKey];
    if (va === null || va === undefined) va = sortAsc ? Infinity : -Infinity;
    if (vb === null || vb === undefined) vb = sortAsc ? Infinity : -Infinity;
    if (typeof va === 'boolean') {{ va = va ? 1 : 0; vb = vb ? 1 : 0; }}
    if (typeof va === 'string') return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
    return sortAsc ? va - vb : vb - va;
  }});
}}

// ─── CSV Export ───
function exportCSV() {{
  const hdr = ['序號','URL','域名','圖片收錄','頁面索引','圖片數量','最佳排名','圖片標題','圖片網址','來源頁面','摘要'];
  const rows = filtered.map(r => [
    r.i, r.url, r.domain,
    r.image_found?'是':'否', r.indexed?'是':'否',
    r.image_count, r.best_position||'',
    r.image_titles, r.image_urls, r.source_urls, r.snippet
  ]);
  const csv = '\\uFEFF' + [hdr,...rows].map(r => r.map(c => `"${{String(c||'').replace(/"/g,'""')}}"`).join(',')).join('\\n');
  const blob = new Blob([csv], {{type:'text/csv;charset=utf-8;'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'image-serp-results.csv';
  a.click();
}}

// ─── Init ───
drawDonut();
drawDomainBars();
renderTable();
</script>
</body>
</html>'''

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✅ 互動式 HTML 報表已儲存: {filepath}")
    print(f"     用瀏覽器開啟即可查看: file:///{os.path.abspath(filepath).replace(os.sep, '/')}")


# ═══════════════════════════════════════════════════════
#  鍵盤控制 (暫停 / 中止)
# ═══════════════════════════════════════════════════════
def keyboard_listener():
    print("  💡 按 [P] 暫停/繼續  |  [Q] 中止查詢\n")
    while not abort_flag.is_set():
        try:
            if sys.platform == "win32":
                import msvcrt
                if msvcrt.kbhit():
                    ch = msvcrt.getch().decode("utf-8", errors="ignore").upper()
                    if ch == "P":
                        if pause_event.is_set():
                            pause_event.clear()
                            print("\n  ⏸  已暫停，按 P 繼續...")
                        else:
                            pause_event.set()
                            print("  ▶  已繼續")
                    elif ch == "Q":
                        abort_flag.set()
                        pause_event.set()
                        print("\n  ⏹  正在中止…")
                        break
            else:
                import select
                if select.select([sys.stdin], [], [], 0.5)[0]:
                    ch = sys.stdin.readline().strip().upper()
                    if ch == "P":
                        if pause_event.is_set():
                            pause_event.clear()
                            print("\n  ⏸  已暫停，輸入 P 繼續...")
                        else:
                            pause_event.set()
                            print("  ▶  已繼續")
                    elif ch == "Q":
                        abort_flag.set()
                        pause_event.set()
                        print("\n  ⏹  正在中止…")
                        break
            time.sleep(0.2)
        except Exception:
            time.sleep(0.5)


# ═══════════════════════════════════════════════════════
#  Worker
# ═══════════════════════════════════════════════════════
def worker(checker: SerpAPIChecker, url: str, delay: float) -> dict:
    pause_event.wait()
    if abort_flag.is_set():
        return {
            "url": url, "domain": domain_of(url), "status": "error",
            "indexed": False, "image_found": False, "image_count": 0,
            "best_position": None, "image_titles": "", "image_urls": "",
            "source_urls": "", "snippet": "使用者中止",
        }
    time.sleep(delay)
    return checker.check(url)


# ═══════════════════════════════════════════════════════
#  主流程
# ═══════════════════════════════════════════════════════
def run(args):
    print()
    print("  ╔═══════════════════════════════════════════════╗")
    print("  ║  🔍 Image SERP 收錄批量檢測 (SerpAPI)        ║")
    print("  ╚═══════════════════════════════════════════════╝")
    print()

    # API Key
    api_key = args.key or os.environ.get("SERPAPI_KEY", "")
    if not api_key:
        print("  ❌ 未提供 SerpAPI Key！")
        print()
        print("  方法 1: 命令列參數")
        print("    python image_serp_checker.py -i urls.txt -o out.xlsx --key YOUR_KEY")
        print()
        print("  方法 2: 環境變數 (設定一次，之後不用重複輸入)")
        print("    set SERPAPI_KEY=YOUR_KEY          (Windows CMD)")
        print("    $env:SERPAPI_KEY='YOUR_KEY'       (PowerShell)")
        print("    export SERPAPI_KEY=YOUR_KEY       (Mac/Linux)")
        print()
        print("  👉 註冊: https://serpapi.com (免費 100 次/月)")
        sys.exit(1)

    # 載入 URL
    urls = load_urls(args.input)
    if not urls:
        print("  ❌ 沒有有效的 URL，請確認輸入檔案格式 (每行一個 http/https 開頭的網址)")
        sys.exit(1)

    credits_per_url = 1 if args.skip_index else 2
    total_credits = len(urls) * credits_per_url

    print(f"  📋 載入 {len(urls)} 個不重複網址")
    print(f"  🔑 API Key: {api_key[:8]}...{api_key[-4:]}")
    print(f"  🌐 搜尋區域: hl={args.hl}  gl={args.gl}")
    print(f"  ⚙️  並行: {args.concurrency}  |  間隔: {args.delay}s  |  {'僅查圖片' if args.skip_index else '圖片 + 索引'}")
    print(f"  💰 預估消耗: ~{total_credits} 次 SerpAPI credit ({credits_per_url}/URL)")
    print(f"  📁 輸出: {args.output}")
    print()

    if not args.yes:
        try:
            ans = input(f"  確定開始查詢？(Y/n) ").strip().lower()
            if ans and ans != "y":
                print("  已取消")
                return
        except (EOFError, KeyboardInterrupt):
            print("\n  已取消")
            return
    print()

    # 初始化 checker
    checker = SerpAPIChecker(
        api_key=api_key, hl=args.hl, gl=args.gl, skip_index=args.skip_index,
    )

    # 驗證 API Key (使用免費的帳號資訊端點，不消耗 credit)
    print("  🔐 驗證 API Key... ", end="", flush=True)
    try:
        acct_url = f"https://serpapi.com/account.json?api_key={api_key}"
        r = requests.get(acct_url, timeout=30)
        r.raise_for_status()
        acct = r.json()

        plan = acct.get("plan_name", "Unknown")
        used = acct.get("this_month_usage", 0)
        limit = acct.get("plan_searches_left", "?")
        total_limit = acct.get("total_searches_left", "?")

        print(f"✓ 有效")
        print(f"  📊 方案: {plan}  |  本月已用: {used} 次  |  剩餘: {total_limit} 次")

        # 檢查額度是否足夠
        if isinstance(total_limit, int) and total_limit < total_credits:
            print(f"\n  ⚠️  剩餘額度 ({total_limit}) 不足以完成全部查詢 ({total_credits} credit)")
            print(f"      將查到額度用完為止，已完成的結果仍會儲存")

    except requests.exceptions.HTTPError as e:
        code = e.response.status_code if e.response else 0
        body = ""
        try:
            body = e.response.text[:300] if e.response else ""
        except Exception:
            pass

        if code == 401 or code == 403:
            print(f"✗ 無效 (HTTP {code})")
            print(f"\n  ❌ API Key 無效或已過期")
            print(f"     請到 https://serpapi.com/manage-api-key 確認")
        elif code == 429:
            print("✗ 額度用盡 (HTTP 429)")
            print("\n  ❌ 本月 API 額度已用完")
        else:
            print(f"✗ HTTP {code}")
            if body:
                print(f"     回應: {body[:200]}")
        sys.exit(1)
    except requests.exceptions.SSLError as e:
        print("✗ SSL 錯誤")
        print(f"\n  ❌ SSL 憑證驗證失敗，可能原因:")
        print(f"     1. 公司防火牆/代理伺服器攔截了 HTTPS 請求")
        print(f"     2. 系統時間不正確")
        print(f"     3. 需要設定 HTTP_PROXY / HTTPS_PROXY 環境變數")
        print(f"\n     詳細錯誤: {str(e)[:200]}")
        sys.exit(1)
    except requests.exceptions.ProxyError as e:
        print("✗ Proxy 錯誤")
        print(f"\n  ❌ 代理伺服器連線失敗")
        print(f"     請檢查 HTTP_PROXY / HTTPS_PROXY 環境變數設定")
        print(f"     詳細錯誤: {str(e)[:200]}")
        sys.exit(1)
    except requests.exceptions.ConnectionError as e:
        print("✗ 連線失敗")
        print(f"\n  ❌ 無法連線到 serpapi.com，可能原因:")
        print(f"     1. 沒有網路連線")
        print(f"     2. DNS 無法解析 serpapi.com")
        print(f"     3. 防火牆封鎖了對外連線")
        print(f"     4. 需要設定 Proxy")
        print(f"\n     測試指令: curl https://serpapi.com/account.json?api_key=test")
        print(f"     詳細錯誤: {str(e)[:200]}")
        sys.exit(1)
    except requests.exceptions.Timeout:
        print("✗ 逾時")
        print(f"\n  ❌ 連線 serpapi.com 逾時 (30 秒)")
        print(f"     請檢查網路連線品質")
        sys.exit(1)
    except Exception as e:
        print(f"✗ 未知錯誤")
        print(f"\n  ❌ {type(e).__name__}: {str(e)[:300]}")
        sys.exit(1)
    print()

    # 鍵盤監聽
    kbd = threading.Thread(target=keyboard_listener, daemon=True)
    kbd.start()

    # 執行批量查詢
    progress = Progress(len(urls))
    results = [None] * len(urls)
    t0 = time.time()

    with ThreadPoolExecutor(max_workers=args.concurrency) as pool:
        futures = {}
        for i, url in enumerate(urls):
            f = pool.submit(worker, checker, url, args.delay)
            futures[f] = i

        for f in as_completed(futures):
            idx = futures[f]
            try:
                r = f.result()
            except Exception as e:
                r = {
                    "url": urls[idx], "domain": domain_of(urls[idx]),
                    "status": "error", "indexed": False, "image_found": False,
                    "image_count": 0, "best_position": None, "image_titles": "",
                    "image_urls": "", "source_urls": "", "snippet": str(e)[:120],
                }
            results[idx] = r
            progress.update(r)

            # API 額度用盡 → 自動中止
            if "額度用盡" in r.get("snippet", ""):
                abort_flag.set()
                pause_event.set()
                print("\n\n  ⚠️  API 額度已用完，自動中止剩餘查詢")
                break

    progress.close()
    abort_flag.set()
    elapsed = time.time() - t0

    results = [r for r in results if r is not None]

    # 統計
    total   = len(results)
    found   = sum(1 for r in results if r["image_found"])
    nfound  = sum(1 for r in results if not r["image_found"] and r["status"] != "error")
    errors  = sum(1 for r in results if r["status"] == "error")
    indexed = sum(1 for r in results if r["indexed"])
    rate    = f"{found / total * 100:.1f}%" if total else "N/A"

    print()
    print("  ┌───────────────────────────────────────┐")
    print("  │          📊 查詢完成統計               │")
    print("  ├───────────────────────────────────────┤")
    print(f"  │  查詢總數       {total:>6}                │")
    print(f"  │  ✅ 圖片已收錄  {found:>6}                │")
    print(f"  │  ❌ 圖片未收錄  {nfound:>6}                │")
    print(f"  │  ⚠️  查詢錯誤    {errors:>6}                │")
    print(f"  │  📑 頁面已索引  {indexed:>6}                │")
    print(f"  │  📈 圖片收錄率  {rate:>6}                │")
    print(f"  │  ⏱️  耗時        {elapsed:>6.1f}s               │")
    print("  └───────────────────────────────────────┘")
    print()

    # 儲存
    out = args.output
    try:
        if out.endswith(".xlsx"):
            print(f"  💾 輸出格式: Excel (.xlsx)")
            save_excel(results, out)
        elif out.endswith(".html") or out.endswith(".htm"):
            print(f"  💾 輸出格式: 互動式 HTML 報表")
            save_html(results, out)
        else:
            if not out.endswith(".csv"):
                out += ".csv"
            print(f"  💾 輸出格式: CSV")
            save_csv(results, out)
    except Exception as e:
        print(f"\n  ❌ 儲存失敗: {type(e).__name__}: {e}")
        # 失敗時自動改存 CSV 作為備援
        fallback = out.rsplit(".", 1)[0] + "_backup.csv"
        try:
            save_csv(results, fallback)
            print(f"  📁 已自動備存 CSV: {fallback}")
        except Exception:
            pass
    print()


# ═══════════════════════════════════════════════════════
#  快速測試 API Key
# ═══════════════════════════════════════════════════════
def test_api_key(args):
    """只測試 API Key 連線，不執行查詢"""
    print()
    print("  🧪 SerpAPI 連線測試")
    print("  " + "─" * 40)

    api_key = args.key or os.environ.get("SERPAPI_KEY", "")
    if not api_key:
        print("  ❌ 未提供 API Key (--key 或環境變數 SERPAPI_KEY)")
        sys.exit(1)

    print(f"  🔑 Key: {api_key[:8]}...{api_key[-4:]}")
    print()

    # 步驟 1: 基本連線
    print("  [1/3] 測試連線到 serpapi.com... ", end="", flush=True)
    try:
        r = requests.get("https://serpapi.com", timeout=15)
        print(f"✓ (HTTP {r.status_code})")
    except requests.exceptions.SSLError:
        print("✗ SSL 錯誤")
        print("        ⮕ 可能是防火牆/代理攔截 HTTPS，或系統時間不正確")
        print("        ⮕ 嘗試: set REQUESTS_CA_BUNDLE=path/to/cert.pem")
        sys.exit(1)
    except requests.exceptions.ConnectionError as e:
        print("✗ 連線失敗")
        print(f"        ⮕ 無法連線 serpapi.com: {str(e)[:150]}")
        print("        ⮕ 請檢查網路 / DNS / 防火牆")
        sys.exit(1)
    except Exception as e:
        print(f"✗ {e}")
        sys.exit(1)

    # 步驟 2: API Key 驗證
    print("  [2/3] 驗證 API Key... ", end="", flush=True)
    try:
        r = requests.get(
            f"https://serpapi.com/account.json?api_key={api_key}",
            timeout=30,
        )
        r.raise_for_status()
        acct = r.json()
        print("✓ 有效")
        print()
        print(f"  📊 帳號資訊:")
        print(f"     方案:     {acct.get('plan_name', 'N/A')}")
        print(f"     本月已用: {acct.get('this_month_usage', 'N/A')} 次")
        print(f"     剩餘額度: {acct.get('total_searches_left', 'N/A')} 次")
        print(f"     信箱:     {acct.get('email', 'N/A')}")
    except requests.exceptions.HTTPError as e:
        code = e.response.status_code if e.response else 0
        print(f"✗ HTTP {code}")
        if code in (401, 403):
            print("        ⮕ API Key 無效，請到 https://serpapi.com/manage-api-key 確認")
        else:
            try:
                print(f"        ⮕ 回應: {e.response.text[:200]}")
            except Exception:
                pass
        sys.exit(1)
    except Exception as e:
        print(f"✗ {type(e).__name__}: {str(e)[:200]}")
        sys.exit(1)

    # 步驟 3: 實際搜尋測試
    print()
    print("  [3/3] 測試搜尋功能... ", end="", flush=True)
    try:
        r = requests.get(SERPAPI_BASE, params={
            "engine": "google_images",
            "q": "test",
            "api_key": api_key,
            "num": 1,
        }, timeout=30)
        r.raise_for_status()
        data = r.json()
        img_count = len(data.get("images_results", []))
        print(f"✓ 成功 (回傳 {img_count} 張圖片)")
    except Exception as e:
        print(f"✗ {type(e).__name__}: {str(e)[:200]}")
        sys.exit(1)

    print()
    print("  ✅ 所有測試通過！可以開始使用：")
    print("     python image_serp_checker.py -i urls.txt -o report.html --key YOUR_KEY")
    print()


# ═══════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════
def main():
    p = argparse.ArgumentParser(
        description="🔍 Image SERP 收錄批量檢測 — 透過 SerpAPI 驗證網頁圖片是否被 Google 圖片搜尋收錄",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
使用範例:

  # 基本用法 (輸出 Excel)
  python image_serp_checker.py -i urls.txt -o results.xlsx --key YOUR_KEY

  # 輸出 CSV
  python image_serp_checker.py -i urls.csv -o results.csv --key YOUR_KEY

  # 輸出互動式 HTML 報表 (可用瀏覽器開啟)
  python image_serp_checker.py -i urls.txt -o report.html --key YOUR_KEY

  # 用環境變數儲存 Key
  set SERPAPI_KEY=YOUR_KEY                                    (CMD)
  python image_serp_checker.py -i urls.txt -o results.xlsx

  # 自訂搜尋區域 (美國英文)
  python image_serp_checker.py -i urls.txt -o out.xlsx --hl en --gl us

  # 只查圖片收錄，省一半 credit
  python image_serp_checker.py -i urls.txt -o out.xlsx --skip-index

  # 加大並行 + 縮短間隔
  python image_serp_checker.py -i urls.txt -o out.xlsx -c 5 -d 0.5

  # 跳過確認直接跑
  python image_serp_checker.py -i urls.txt -o out.xlsx -y

  # 只測試 API Key 連線 (不消耗 credit)
  python image_serp_checker.py --test --key YOUR_KEY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """,
    )

    p.add_argument("-i", "--input", required=False,
                   help="輸入檔案 (.txt / .csv，每行一個 URL)")
    p.add_argument("-o", "--output", required=False,
                   help="輸出檔案 (.xlsx / .csv / .html)")
    p.add_argument("--key",
                   help="SerpAPI Key (或環境變數 SERPAPI_KEY)")
    p.add_argument("-c", "--concurrency", type=int, default=DEFAULT_CONCURRENCY,
                   help=f"並行查詢數 (預設: {DEFAULT_CONCURRENCY})")
    p.add_argument("-d", "--delay", type=float, default=DEFAULT_DELAY,
                   help=f"查詢間隔秒數 (預設: {DEFAULT_DELAY})")
    p.add_argument("--hl", default="zh-TW",
                   help="搜尋語言 (預設: zh-TW)")
    p.add_argument("--gl", default="tw",
                   help="搜尋國家/地區 (預設: tw)")
    p.add_argument("--skip-index", action="store_true",
                   help="跳過頁面索引檢查，只查圖片收錄 (省一半 credit)")
    p.add_argument("-y", "--yes", action="store_true",
                   help="跳過確認，直接開始")
    p.add_argument("--test", action="store_true",
                   help="只測試 API Key 連線，不執行查詢")

    args = p.parse_args()

    if args.test:
        test_api_key(args)
    else:
        if not args.input or not args.output:
            p.error("需要提供 -i (輸入檔案) 和 -o (輸出檔案)，或使用 --test 只測試連線")
        run(args)


if __name__ == "__main__":
    main()